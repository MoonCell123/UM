"""
Publication-quality visualization for UVM D3/M3 binary classification benchmark.
Three scientific questions with figures and three-line tables.
"""

import pandas as pd
import numpy as np
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
from matplotlib.gridspec import GridSpec
from matplotlib.lines import Line2D
import warnings
import os
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill

warnings.filterwarnings("ignore")
matplotlib.rcParams.update({
    "font.family": "Arial",
    "axes.spines.top": False,
    "axes.spines.right": False,
    "pdf.fonttype": 42,
    "ps.fonttype": 42,
})

# ── paths ──────────────────────────────────────────────────────────────────────
RESULT_DIR = os.path.join(os.path.dirname(__file__),
                          "benchmark_output", "20260331_171049")
OUT_DIR = os.path.join(RESULT_DIR, "figures")
os.makedirs(OUT_DIR, exist_ok=True)

# ── load data ─────────────────────────────────────────────────────────────────
df = pd.read_csv(os.path.join(RESULT_DIR, "summary.csv"))
ga = pd.read_csv(os.path.join(RESULT_DIR, "group_analysis.csv"))

# ── model metadata ─────────────────────────────────────────────────────────────
TCGA_MODELS = {
    "kaiko-vitb16", "lunit-vits8", "kaiko-vits8", "kaiko-vitl14",
    "phikon_v2", "kaiko-vits16", "kaiko-vitb8", "midnight12k",
}
DISPLAY_NAMES = {
    "hoptimus1": "H-optimus-1",
    "hoptimus0": "H-optimus-0",
    "uni_v1": "UNI v1",
    "uni_v2": "UNI v2",
    "conch_v1": "CONCH v1",
    "conch_v15": "CONCH v1.5",
    "virchow": "Virchow",
    "virchow2": "Virchow2",
    "gigapath": "GigaPath",
    "musk": "MUSK",
    "hibou_l": "Hibou-L",
    "phikon": "Phikon",
    "phikon_v2": "Phikon-v2",
    "kaiko-vits8": "Kaiko ViT-S/8",
    "kaiko-vits16": "Kaiko ViT-S/16",
    "kaiko-vitb8": "Kaiko ViT-B/8",
    "kaiko-vitb16": "Kaiko ViT-B/16",
    "kaiko-vitl14": "Kaiko ViT-L/14",
    "lunit-vits8": "LUNIT ViT-S/8",
    "midnight12k": "Midnight-12k",
    "resnet50": "ResNet-50",
}

# add group column
df["group"] = df["model"].apply(
    lambda m: "TCGA-Pretrained" if m in TCGA_MODELS else
    ("Baseline" if m == "resnet50" else "Private-Pretrained")
)
df["display"] = df["model"].map(DISPLAY_NAMES).fillna(df["model"])
df_sorted = df.sort_values("auc_macro_mean", ascending=False).reset_index(drop=True)
df_sorted["rank"] = range(1, len(df_sorted) + 1)

# ── color palette ─────────────────────────────────────────────────────────────
C_PRIVATE  = "#2E86AB"   # deep blue  – non-TCGA foundation
C_TCGA     = "#E07B39"   # warm amber – TCGA-pretrained
C_BASELINE = "#888888"   # neutral gray – ResNet-50
C_KAIKO    = "#9B59B6"   # purple – kaiko highlight

ALPHA_BAR  = 0.85
DPI        = 300
FIG_W      = 10  # inches, single-column ≈ 3.5, double-column ≈ 7, wide ≈ 10


# ══════════════════════════════════════════════════════════════════════════════
# HELPER: three-line table as PNG
# ══════════════════════════════════════════════════════════════════════════════
def save_three_line_table(table_df, title, filename,
                          col_widths=None, fontsize=8):
    n_rows, n_cols = table_df.shape
    row_h = 0.32
    header_h = 0.45
    fig_h = header_h + n_rows * row_h + 0.3
    fig_w = sum(col_widths) if col_widths else n_cols * 1.5

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.set_xlim(0, fig_w)
    ax.set_ylim(0, fig_h)
    ax.axis("off")

    # column widths → x positions
    if col_widths is None:
        col_widths = [fig_w / n_cols] * n_cols
    x_pos = [sum(col_widths[:i]) for i in range(n_cols)]

    top_y    = fig_h - 0.05
    header_y = fig_h - header_h + 0.05
    data_top = header_y - 0.08
    bottom_y = 0.15

    lw = 1.2
    # top rule
    ax.axhline(top_y, color="black", linewidth=lw * 1.5,
               xmin=0, xmax=1)
    # mid rule (below header)
    ax.axhline(header_y, color="black", linewidth=lw,
               xmin=0, xmax=1)
    # bottom rule
    ax.axhline(bottom_y, color="black", linewidth=lw * 1.5,
               xmin=0, xmax=1)

    # header
    for ci, (col, cw, xp) in enumerate(zip(table_df.columns, col_widths, x_pos)):
        align = "left" if ci == 0 else "center"
        ax.text(xp + (0.05 if ci == 0 else cw / 2),
                (top_y + header_y) / 2,
                col, ha=align, va="center",
                fontsize=fontsize, fontweight="bold",
                fontfamily="Arial")

    # data rows – alternating background
    for ri, row in enumerate(table_df.itertuples(index=False)):
        y = data_top - ri * row_h
        if ri % 2 == 1:
            rect = mpatches.FancyBboxPatch(
                (0, y - row_h * 0.85), fig_w, row_h * 0.9,
                boxstyle="square,pad=0", linewidth=0,
                facecolor="#f5f5f5")
            ax.add_patch(rect)
        for ci, (val, cw, xp) in enumerate(zip(row, col_widths, x_pos)):
            align = "left" if ci == 0 else "center"
            color = "black"
            fw = "normal"
            style = "normal"
            # highlight resnet50 row
            if str(val).startswith("ResNet"):
                color = "#555555"
                style = "italic"
            ax.text(xp + (0.05 if ci == 0 else cw / 2),
                    y - row_h * 0.35,
                    str(val), ha=align, va="center",
                    fontsize=fontsize, color=color,
                    fontweight=fw, fontstyle=style,
                    fontfamily="Arial")

    ax.text(fig_w / 2, top_y + 0.04, title,
            ha="center", va="bottom",
            fontsize=fontsize + 1, fontweight="bold", fontfamily="Arial")
    ax.set_ylim(-0.05, top_y + 0.18)
    plt.tight_layout(pad=0.1)
    fig.savefig(os.path.join(OUT_DIR, filename),
                dpi=DPI, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    print(f"  Saved: {filename}")

    # ── also export to Excel (.xlsx) ──────────────────────────────────────────
    xlsx_name = filename.replace(".png", ".xlsx")
    _save_excel_table(table_df, title, os.path.join(OUT_DIR, xlsx_name))


def _save_excel_table(tdf, title, xlsx_path):
    """Save a DataFrame as a three-line-table formatted Excel workbook."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Table"

    thick = Side(style="medium", color="000000")
    thin  = Side(style="thin",   color="000000")
    none  = Side(style=None)

    top_border    = Border(top=thick, bottom=thin, left=none, right=none)
    mid_border    = Border(top=thin,  bottom=none, left=none, right=none)
    bot_border    = Border(top=none,  bottom=thick, left=none, right=none)
    data_border   = Border(top=none,  bottom=none,  left=none, right=none)
    fill_alt      = PatternFill("solid", fgColor="F5F5F5")

    # title row
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(tdf.columns))
    tc = ws.cell(row=1, column=1, value=title)
    tc.font      = Font(name="Arial", bold=True, size=11)
    tc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    tc.border    = Border(top=thick, left=none, right=none, bottom=none)
    ws.row_dimensions[1].height = 28

    # header row
    for ci, col in enumerate(tdf.columns, start=1):
        cell = ws.cell(row=2, column=ci, value=col)
        cell.font      = Font(name="Arial", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center" if ci > 1 else "left",
                                   vertical="center", wrap_text=True)
        cell.border    = top_border
    ws.row_dimensions[2].height = 20

    # last header row bottom rule
    for ci in range(1, len(tdf.columns)+1):
        ws.cell(row=2, column=ci).border = Border(
            top=thick, bottom=thin, left=none, right=none)

    # data rows
    n_rows = len(tdf)
    for ri, row in enumerate(tdf.itertuples(index=False), start=3):
        is_last = (ri == n_rows + 2)
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=str(val))
            cell.font      = Font(name="Arial", size=9.5,
                                  italic=str(val).startswith("ResNet"))
            cell.alignment = Alignment(horizontal="center" if ci > 1 else "left",
                                       vertical="center", wrap_text=True)
            cell.border    = bot_border if is_last else data_border
        if ri % 2 == 0:
            for ci in range(1, len(tdf.columns)+1):
                ws.cell(row=ri, column=ci).fill = fill_alt
        ws.row_dimensions[ri].height = 16

    # column widths (auto-fit approximation)
    for ci, col in enumerate(tdf.columns, start=1):
        max_len = max([len(str(col))] +
                      [len(str(v)) for v in tdf.iloc[:, ci-1]])
        ws.column_dimensions[
            openpyxl.utils.get_column_letter(ci)
        ].width = min(max_len * 1.35 + 2, 40)

    wb.save(xlsx_path)
    print(f"  Saved: {os.path.basename(xlsx_path)}")


# ══════════════════════════════════════════════════════════════════════════════
# TABLE 0 – Full ranking (三线表)
# ══════════════════════════════════════════════════════════════════════════════
def make_full_ranking_table():
    rows = []
    for _, r in df_sorted.iterrows():
        group_label = {
            "TCGA-Pretrained": "TCGA",
            "Private-Pretrained": "Private",
            "Baseline": "Baseline",
        }[r["group"]]
        rows.append({
            "Rank": int(r["rank"]),
            "Model": r["display"],
            "Pretraining Data": group_label,
            "AUC (mean±std)": f"{r['auc_macro_mean']:.4f}±{r['auc_macro_std']:.4f}",
            "Accuracy": f"{r['acc_mean']:.4f}",
            "F1 (weighted)": f"{r['f1_weighted_mean']:.4f}",
        })
    tdf = pd.DataFrame(rows)
    save_three_line_table(
        tdf,
        title="Table 1. Complete Benchmark Ranking – UVM D3/M3 Binary Classification (n=80)",
        filename="table3_full_ranking.png",
        col_widths=[0.55, 2.0, 1.5, 1.8, 1.2, 1.6],
        fontsize=8,
    )


# ══════════════════════════════════════════════════════════════════════════════
# FIGURE 1 + TABLE Q1 – Transfer value
# ══════════════════════════════════════════════════════════════════════════════
def make_q1():
    # ── figure ────────────────────────────────────────────────────────────────
    fig, ax = plt.subplots(figsize=(FIG_W, 5.2))

    bar_colors = []
    for _, r in df_sorted.iterrows():
        if r["group"] == "Baseline":
            bar_colors.append(C_BASELINE)
        elif r["group"] == "TCGA-Pretrained":
            bar_colors.append(C_TCGA)
        else:
            bar_colors.append(C_PRIVATE)

    x = np.arange(len(df_sorted))
    bars = ax.bar(x, df_sorted["auc_macro_mean"],
                  color=bar_colors, alpha=ALPHA_BAR,
                  width=0.65, zorder=3, edgecolor="white", linewidth=0.4)

    # error bars
    ax.errorbar(x, df_sorted["auc_macro_mean"],
                yerr=df_sorted["auc_macro_std"],
                fmt="none", color="black", capsize=3,
                capthick=0.8, linewidth=0.8, zorder=4)

    # ResNet50 baseline dashed line
    resnet_auc = df_sorted.loc[df_sorted["model"] == "resnet50",
                               "auc_macro_mean"].values[0]
    ax.axhline(resnet_auc, color=C_BASELINE, linewidth=1.4,
               linestyle="--", zorder=2, label=f"ResNet-50 baseline (AUC={resnet_auc:.4f})")

    # annotation: p-value from Q1 (sign test / binomial, primary test)
    q1 = ga[ga["question"] == "Q1_transfer_value"].iloc[0]
    p_sign = float(q1["p_value_sign"])
    sig_star = "***" if p_sign < 0.001 else ("**" if p_sign < 0.01 else ("*" if p_sign < 0.05 else "n.s."))
    n_above = int(q1["n_above_baseline"])
    n_fm = len(df_sorted[df_sorted["model"] != "resnet50"])
    p_txt = (f"Sign test / binomial (one-sided)\n"
             f"{n_above}/{n_fm} foundation models > ResNet-50\n"
             f"p = {p_sign:.2e}  {sig_star}")
    ax.text(0.98, 0.04, p_txt, transform=ax.transAxes,
            ha="right", va="bottom", fontsize=8,
            bbox=dict(boxstyle="round,pad=0.35", facecolor="white",
                      edgecolor="#cccccc", linewidth=0.8))

    # axis labels
    ax.set_xticks(x)
    ax.set_xticklabels(df_sorted["display"], rotation=40, ha="right",
                       fontsize=8)
    ax.set_ylabel("AUC (5-fold CV, mean ± std)", fontsize=10)
    ax.set_title(
        "Q1  Transfer Value of Pathology Foundation Models vs. ResNet-50\n"
        "UVM D3/M3 Classification  |  n = 80  |  ABMIL",
        fontsize=10, fontweight="bold", pad=10)

    ax.set_ylim(0.60, 1.02)
    ax.yaxis.grid(True, color="#e0e0e0", linewidth=0.7, zorder=0)
    ax.set_axisbelow(True)

    legend_handles = [
        mpatches.Patch(facecolor=C_PRIVATE,  alpha=ALPHA_BAR, label="Private-Pretrained Foundation"),
        mpatches.Patch(facecolor=C_TCGA,     alpha=ALPHA_BAR, label="TCGA-Pretrained Foundation"),
        mpatches.Patch(facecolor=C_BASELINE, alpha=ALPHA_BAR, label="Baseline (ResNet-50)"),
    ]
    ax.legend(handles=legend_handles, loc="upper right",
              fontsize=8, frameon=True, framealpha=0.9,
              edgecolor="#cccccc")

    plt.tight_layout()
    fig.savefig(os.path.join(OUT_DIR, "fig3_q1_transfer_value.pdf"),
                dpi=DPI, bbox_inches="tight", facecolor="white")
    fig.savefig(os.path.join(OUT_DIR, "fig3_q1_transfer_value.png"),
                dpi=DPI, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    print("  Saved: fig3_q1_transfer_value.pdf/.png")

    # ── table ──────────────────────────────────────────────────────────────────
    rows = []
    for _, r in df_sorted[df_sorted["model"] != "resnet50"].iterrows():
        delta = r["auc_macro_mean"] - resnet_auc
        rows.append({
            "Model": r["display"],
            "Pretraining": r["group"].replace("-Pretrained", ""),
            "AUC (mean±std)": f"{r['auc_macro_mean']:.4f}±{r['auc_macro_std']:.4f}",
            "ΔAUC vs ResNet-50": f"{delta:+.4f}",
        })
    rows.append({
        "Model": "ResNet-50 (baseline)",
        "Pretraining": "ImageNet",
        "AUC (mean±std)": f"{resnet_auc:.4f}±{df_sorted.loc[df_sorted['model']=='resnet50','auc_macro_std'].values[0]:.4f}",
        "ΔAUC vs ResNet-50": "—",
    })
    tdf = pd.DataFrame(rows)
    save_three_line_table(
        tdf,
        title=(f"Table Q1. Foundation Models vs. ResNet-50  "
               f"(sign test: p={p_sign:.2e} {sig_star}, {n_above}/{n_fm} > baseline)"),
        filename="table4_q1_transfer_value.png",
        col_widths=[2.2, 1.4, 1.9, 1.7],
        fontsize=8,
    )


# ══════════════════════════════════════════════════════════════════════════════
# FIGURE 2 + TABLE Q2 – TCGA overlap effect
# ══════════════════════════════════════════════════════════════════════════════
def make_q2():
    foundation = df_sorted[df_sorted["model"] != "resnet50"].copy()
    tcga_aucs    = foundation[foundation["group"] == "TCGA-Pretrained"]["auc_macro_mean"].values
    private_aucs = foundation[foundation["group"] == "Private-Pretrained"]["auc_macro_mean"].values

    q2 = ga[ga["question"] == "Q2_contamination_effect"].iloc[0]

    fig, axes = plt.subplots(1, 2, figsize=(FIG_W, 5.0),
                             gridspec_kw={"width_ratios": [1.6, 1]})

    # ── left: strip + box ─────────────────────────────────────────────────────
    ax = axes[0]
    groups_data = [tcga_aucs, private_aucs]
    positions   = [1, 2]
    colors      = [C_TCGA, C_PRIVATE]
    labels      = ["TCGA-Pretrained\n(n=8)", "Private-Pretrained\n(n=12)"]

    bp = ax.boxplot(groups_data, positions=positions, widths=0.38,
                    patch_artist=True, notch=False,
                    medianprops=dict(color="white", linewidth=2),
                    whiskerprops=dict(linewidth=1.2),
                    capprops=dict(linewidth=1.2),
                    flierprops=dict(marker="", linewidth=0))
    for patch, color in zip(bp["boxes"], colors):
        patch.set_facecolor(color)
        patch.set_alpha(0.5)
        patch.set_linewidth(1.2)

    # jitter strip
    rng = np.random.default_rng(42)
    for aucs, pos, color in zip(groups_data, positions, colors):
        jitter = rng.uniform(-0.1, 0.1, len(aucs))
        ax.scatter(pos + jitter, aucs, color=color, s=42,
                   zorder=5, edgecolors="white", linewidths=0.5, alpha=0.9)

    # significance bracket
    y_max = max(np.max(tcga_aucs), np.max(private_aucs))
    y_sig = y_max + 0.025
    ax.plot([1, 1, 2, 2], [y_sig - 0.01, y_sig, y_sig, y_sig - 0.01],
            color="black", linewidth=1.2)
    p_perm = float(q2['p_value_permutation'])
    sig_q2 = "***" if p_perm < 0.001 else ("**" if p_perm < 0.01 else ("*" if p_perm < 0.05 else "n.s."))
    p_label = f"p = {p_perm:.4f} {sig_q2}"
    ax.text(1.5, y_sig + 0.005, p_label, ha="center", va="bottom",
            fontsize=9, fontweight="bold")

    ax.set_xticks(positions)
    ax.set_xticklabels(labels, fontsize=9)
    ax.set_ylabel("AUC (5-fold CV)", fontsize=10)
    ax.set_title("Group Comparison: TCGA vs. Private Pretraining",
                 fontsize=10, fontweight="bold")
    ax.set_ylim(0.74, y_sig + 0.055)
    ax.yaxis.grid(True, color="#e0e0e0", linewidth=0.7)
    ax.set_axisbelow(True)

    # means text
    for aucs, pos, color in zip(groups_data, positions, colors):
        ax.text(pos, np.mean(aucs) - 0.013,
                f"mean={np.mean(aucs):.4f}",
                ha="center", va="top", fontsize=7.5, color=color,
                fontweight="bold")

    # ── right: ranked dot plot by group ───────────────────────────────────────
    ax2 = axes[1]
    fnd = foundation.copy().reset_index(drop=True)
    y_vals = np.arange(len(fnd))

    for i, row in fnd.iterrows():
        color = C_TCGA if row["group"] == "TCGA-Pretrained" else C_PRIVATE
        ax2.errorbar(row["auc_macro_mean"], i,
                     xerr=row["auc_macro_std"],
                     fmt="o", color=color, markersize=5,
                     capsize=2.5, capthick=0.8, linewidth=0.8,
                     ecolor=color, alpha=0.85)

    ax2.set_yticks(y_vals)
    ax2.set_yticklabels(fnd["display"], fontsize=7.5)
    ax2.set_xlabel("AUC (mean ± std)", fontsize=9)
    ax2.set_title("Individual Model AUCs", fontsize=9, fontweight="bold")
    ax2.axvline(resnet_auc := df_sorted.loc[df_sorted["model"]=="resnet50",
                                            "auc_macro_mean"].values[0],
                color=C_BASELINE, linewidth=1.2, linestyle="--",
                label=f"ResNet-50 ({resnet_auc:.4f})")
    ax2.legend(fontsize=7.5, frameon=True, framealpha=0.9,
               edgecolor="#cccccc", loc="lower right")
    ax2.xaxis.grid(True, color="#e0e0e0", linewidth=0.7)
    ax2.set_axisbelow(True)

    legend_handles = [
        mpatches.Patch(facecolor=C_TCGA,    alpha=0.7, label="TCGA-Pretrained"),
        mpatches.Patch(facecolor=C_PRIVATE, alpha=0.7, label="Private-Pretrained"),
    ]
    axes[0].legend(handles=legend_handles, loc="lower right",
                   fontsize=8, frameon=True, framealpha=0.9, edgecolor="#cccccc")

    fig.suptitle(
        "Q2  Effect of Slide-Level TCGA Pretraining Overlap on UVM Classification Performance",
        fontsize=10.5, fontweight="bold", y=1.01)
    plt.tight_layout()
    fig.savefig(os.path.join(OUT_DIR, "fig6_q2_tcga_overlap.pdf"),
                dpi=DPI, bbox_inches="tight", facecolor="white")
    fig.savefig(os.path.join(OUT_DIR, "fig6_q2_tcga_overlap.png"),
                dpi=DPI, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    print("  Saved: fig6_q2_tcga_overlap.pdf/.png")

    # ── table ──────────────────────────────────────────────────────────────────
    rows = []
    for g_label, g_key, g_df in [
        ("TCGA-Pretrained", "TCGA-Pretrained",  foundation[foundation["group"]=="TCGA-Pretrained"]),
        ("Private-Pretrained", "Private-Pretrained", foundation[foundation["group"]=="Private-Pretrained"]),
    ]:
        for _, r in g_df.iterrows():
            rows.append({
                "Pretraining Data": g_label,
                "Model": r["display"],
                "AUC (mean±std)": f"{r['auc_macro_mean']:.4f}±{r['auc_macro_std']:.4f}",
                "Accuracy": f"{r['acc_mean']:.4f}",
            })
    tdf = pd.DataFrame(rows)
    save_three_line_table(
        tdf,
        title=(f"Table Q2. TCGA-Pretrained vs. Private-Pretrained  "
               f"(ΔAUC={q2['delta_AUC']:+.4f}, p={p_perm:.4f}, Permutation test two-sided)"),
        filename="table5_q2_tcga_overlap.png",
        col_widths=[1.9, 2.0, 2.0, 1.3],
        fontsize=8,
    )


# ══════════════════════════════════════════════════════════════════════════════
# FIGURE 3 + TABLE Q3 – Kaiko 2×2 factorial
# ══════════════════════════════════════════════════════════════════════════════
def make_q3():
    kaiko_data = {
        ("ViT-S", "Patch 16"): ("kaiko-vits16", 0.8393, 0.0943),
        ("ViT-S", "Patch 8"):  ("kaiko-vits8",  0.8713, 0.0421),
        ("ViT-B", "Patch 16"): ("kaiko-vitb16", 0.8805, 0.0699),
        ("ViT-B", "Patch 8"):  ("kaiko-vitb8",  0.8300, 0.0688),
    }

    sizes   = ["ViT-S", "ViT-B"]
    patches = ["Patch 8", "Patch 16"]

    # values matrix
    vals = np.array([
        [kaiko_data[("ViT-S", "Patch 8")][1],  kaiko_data[("ViT-S", "Patch 16")][1]],
        [kaiko_data[("ViT-B", "Patch 8")][1],  kaiko_data[("ViT-B", "Patch 16")][1]],
    ])
    stds = np.array([
        [kaiko_data[("ViT-S", "Patch 8")][2],  kaiko_data[("ViT-S", "Patch 16")][2]],
        [kaiko_data[("ViT-B", "Patch 8")][2],  kaiko_data[("ViT-B", "Patch 16")][2]],
    ])

    fig, axes = plt.subplots(1, 2, figsize=(FIG_W, 4.5),
                             gridspec_kw={"width_ratios": [1, 1.1]})

    # ── left: interaction plot ─────────────────────────────────────────────────
    ax = axes[0]
    line_colors = ["#E07B39", "#2E86AB"]   # patch8=amber, patch16=blue
    markers     = ["o", "s"]

    for pi, (patch_lbl, color, marker) in enumerate(
            zip(patches, line_colors, markers)):
        y = [vals[0, pi], vals[1, pi]]
        e = [stds[0, pi], stds[1, pi]]
        ax.errorbar([0, 1], y, yerr=e,
                    color=color, marker=marker, markersize=8,
                    linewidth=1.8, capsize=4, capthick=1,
                    label=patch_lbl, zorder=4)
        for xi, (yi, ei) in enumerate(zip(y, e)):
            ax.text(xi + 0.04, yi + 0.002,
                    f"{yi:.4f}", fontsize=8, color=color, va="bottom")

    ax.set_xticks([0, 1])
    ax.set_xticklabels(["ViT-S\n(~22M params)", "ViT-B\n(~86M params)"],
                       fontsize=9)
    ax.set_ylabel("AUC (5-fold CV, mean ± std)", fontsize=10)
    ax.set_ylim(0.78, 0.925)
    ax.yaxis.grid(True, color="#e0e0e0", linewidth=0.7)
    ax.set_axisbelow(True)
    ax.set_title("Interaction Plot\n(Model Capacity × Patch Size)", fontsize=9.5,
                 fontweight="bold")
    ax.legend(fontsize=8.5, frameon=True, framealpha=0.9, edgecolor="#cccccc",
              title="Patch Size", title_fontsize=8)

    # main effect annotations
    ax.annotate("", xy=(1, np.mean(vals[1, :])),
                xytext=(0, np.mean(vals[0, :])),
                arrowprops=dict(arrowstyle="->", color="gray",
                                linewidth=1.2, linestyle="dashed"))

    # ── right: heatmap ────────────────────────────────────────────────────────
    ax2 = axes[1]
    im = ax2.imshow(vals, cmap="Blues", vmin=0.80, vmax=0.91,
                    aspect="auto")

    for i in range(2):
        for j in range(2):
            val = vals[i, j]
            std = stds[i, j]
            txt = f"{val:.4f}\n±{std:.4f}"
            text_color = "white" if val > 0.865 else "black"
            ax2.text(j, i, txt, ha="center", va="center",
                     fontsize=9.5, color=text_color, fontweight="bold")

    ax2.set_xticks([0, 1])
    ax2.set_xticklabels(["Patch 8\n(high-res)", "Patch 16\n(low-res)"], fontsize=9)
    ax2.set_yticks([0, 1])
    ax2.set_yticklabels(["ViT-S\n(~22M)", "ViT-B\n(~86M)"], fontsize=9)
    ax2.set_xlabel("Spatial Resolution (patch size)", fontsize=9)
    ax2.set_ylabel("Model Capacity", fontsize=9)
    ax2.set_title("AUC Heatmap\n(2×2 Factorial: Kaiko Series)", fontsize=9.5,
                  fontweight="bold")

    cbar = fig.colorbar(im, ax=ax2, shrink=0.75, pad=0.03)
    cbar.set_label("AUC", fontsize=8)
    cbar.ax.tick_params(labelsize=7.5)

    fig.suptitle(
        "Q3  Model Capacity × Patch Size Factorial Design — Kaiko Series\n"
        "Fixed: TCGA Pretraining · DINO Algorithm · Slide-Level Overlap",
        fontsize=10.5, fontweight="bold", y=1.01)
    plt.tight_layout()
    fig.savefig(os.path.join(OUT_DIR, "fig8_q3_kaiko_factorial.pdf"),
                dpi=DPI, bbox_inches="tight", facecolor="white")
    fig.savefig(os.path.join(OUT_DIR, "fig8_q3_kaiko_factorial.png"),
                dpi=DPI, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    print("  Saved: fig8_q3_kaiko_factorial.pdf/.png")

    # ── table ──────────────────────────────────────────────────────────────────
    rows = [
        {"Model": "Kaiko ViT-B/8",  "Capacity": "ViT-B (~86M)", "Patch Size": "8 (high-res)",
         "AUC (mean±std)": "0.8300±0.0688", "vs ViT-B/16": "−0.0505", "vs ViT-S/8": "−0.0413"},
        {"Model": "Kaiko ViT-B/16", "Capacity": "ViT-B (~86M)", "Patch Size": "16 (low-res)",
         "AUC (mean±std)": "0.8805±0.0699", "vs ViT-B/16": "—",      "vs ViT-S/8": "+0.0092"},
        {"Model": "Kaiko ViT-S/8",  "Capacity": "ViT-S (~22M)", "Patch Size": "8 (high-res)",
         "AUC (mean±std)": "0.8713±0.0421", "vs ViT-B/16": "−0.0092", "vs ViT-S/8": "—"},
        {"Model": "Kaiko ViT-S/16", "Capacity": "ViT-S (~22M)", "Patch Size": "16 (low-res)",
         "AUC (mean±std)": "0.8393±0.0943", "vs ViT-B/16": "−0.0412", "vs ViT-S/8": "−0.0320"},
    ]
    tdf = pd.DataFrame(rows)
    save_three_line_table(
        tdf,
        title="Table Q3. Kaiko 2×2 Factorial Design — Model Capacity × Spatial Resolution",
        filename="table6_q3_kaiko_factorial.png",
        col_widths=[1.8, 1.6, 1.6, 2.0, 1.3, 1.3],
        fontsize=8,
    )


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print(f"Output directory: {OUT_DIR}\n")

    print("[Table 0] Full ranking table ...")
    make_full_ranking_table()

    print("[Q1] Transfer value figure + table ...")
    make_q1()

    print("[Q2] TCGA overlap figure + table ...")
    make_q2()

    print("[Q3] Kaiko factorial figure + table ...")
    make_q3()

    print("\nDone. All figures saved to:", OUT_DIR)
