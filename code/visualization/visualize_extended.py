"""
Extended publication-quality figures for UVM D3/M3 benchmark.
Covers: clinical table, pipeline schematic, model overview,
        per-fold boxplots, ROC curves, scatter plots, metric correlation.
"""

import os, warnings
import numpy as np
import pandas as pd
import openpyxl
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
import matplotlib
import matplotlib.pyplot as plt
import matplotlib.patches as mpatches
import matplotlib.patheffects as pe
from matplotlib.patches import FancyArrowPatch, FancyBboxPatch
from matplotlib.lines import Line2D
from matplotlib.gridspec import GridSpec
from sklearn.metrics import roc_curve, auc as sk_auc
from scipy import stats
import itertools

warnings.filterwarnings("ignore")
matplotlib.rcParams.update({
    "font.family": "Arial",
    "axes.spines.top": False,
    "axes.spines.right": False,
    "pdf.fonttype": 42, "ps.fonttype": 42,
    "axes.labelsize": 10, "xtick.labelsize": 8.5, "ytick.labelsize": 8.5,
})

# ── paths ──────────────────────────────────────────────────────────────────────
BASE     = os.path.dirname(__file__)
RES_DIR  = os.path.join(BASE, "benchmark_output", "20260331_171049")
OUT_DIR  = os.path.join(RES_DIR, "figures")
CLIN_CSV = "D:/Datas of lab/UVM/临床表/临床信息表.csv"
os.makedirs(OUT_DIR, exist_ok=True)

DPI = 300

# ── palette ───────────────────────────────────────────────────────────────────
C_PRIVATE  = "#2E86AB"
C_TCGA     = "#E07B39"
C_BASELINE = "#888888"
C_D3       = "#4C9BE8"
C_M3       = "#E85C4C"
ALPHA      = 0.82

# ── display names ─────────────────────────────────────────────────────────────
DISP = {
    "hoptimus1":"H-optimus-1","hoptimus0":"H-optimus-0",
    "uni_v1":"UNI v1","uni_v2":"UNI v2",
    "conch_v1":"CONCH v1","conch_v15":"CONCH v1.5",
    "virchow":"Virchow","virchow2":"Virchow2",
    "gigapath":"GigaPath","musk":"MUSK","hibou_l":"Hibou-L",
    "phikon":"Phikon","phikon_v2":"Phikon-v2",
    "kaiko-vits8":"Kaiko ViT-S/8","kaiko-vits16":"Kaiko ViT-S/16",
    "kaiko-vitb8":"Kaiko ViT-B/8","kaiko-vitb16":"Kaiko ViT-B/16",
    "kaiko-vitl14":"Kaiko ViT-L/14",
    "lunit-vits8":"LUNIT ViT-S/8","midnight12k":"Midnight-12k",
    "resnet50":"ResNet-50",
}

TCGA_SET = {"kaiko-vitb16","lunit-vits8","kaiko-vits8","kaiko-vitl14",
            "phikon_v2","kaiko-vits16","kaiko-vitb8","midnight12k"}

# ── model metadata ─────────────────────────────────────────────────────────────
# (display_name, institution, paradigm, architecture, approx_params_M, D_feat)
MODEL_META = {
    "hoptimus1":   ("H-optimus-1", "Owkin",     "SSL",      "ViT-G", 1100, 1536),
    "hoptimus0":   ("H-optimus-0", "Owkin",     "SSL",      "ViT-G", 1100, 1536),
    "uni_v1":      ("UNI v1",      "Harvard",   "SSL",      "ViT-L",  307, 1024),
    "uni_v2":      ("UNI v2",      "Harvard",   "SSL",      "ViT-G", 1100, 1536),
    "conch_v1":    ("CONCH v1",    "Harvard",   "VLP",      "ViT-B",   87,  512),
    "conch_v15":   ("CONCH v1.5",  "Harvard",   "VLP",      "ViT-B",   87,  768),
    "virchow":     ("Virchow",     "Microsoft", "SSL",      "ViT-H",  632, 2560),
    "virchow2":    ("Virchow2",    "Microsoft", "SSL",      "ViT-H",  632, 2560),
    "gigapath":    ("GigaPath",    "Microsoft", "SSL",      "ViT-G", 1100, 1536),
    "musk":        ("MUSK",        "Tsinghua",  "VLP",      "ViT-L",  307, 1024),
    "hibou_l":     ("Hibou-L",     "HistAI",    "SSL",      "ViT-L",  307, 1024),
    "phikon":      ("Phikon",      "Owkin",     "SSL",      "ViT-B",   87,  768),
    "phikon_v2":   ("Phikon-v2",   "Owkin",     "SSL",      "ViT-L",  307, 1024),
    "kaiko-vits8": ("Kaiko ViT-S/8","Kaiko.ai", "SSL",      "ViT-S",   22,  384),
    "kaiko-vits16":("Kaiko ViT-S/16","Kaiko.ai","SSL",      "ViT-S",   22,  384),
    "kaiko-vitb8": ("Kaiko ViT-B/8","Kaiko.ai", "SSL",      "ViT-B",   87,  768),
    "kaiko-vitb16":("Kaiko ViT-B/16","Kaiko.ai","SSL",      "ViT-B",   87,  768),
    "kaiko-vitl14":("Kaiko ViT-L/14","Kaiko.ai","SSL",      "ViT-L",  307, 1024),
    "lunit-vits8": ("LUNIT ViT-S/8","Lunit",    "SSL",      "ViT-S",   22,  384),
    "midnight12k": ("Midnight-12k","AIGC Lab",  "SSL",      "ViT-G", 1100, 1536),
    "resnet50":    ("ResNet-50",   "PyTorch",   "Supervised","ResNet-50",25, 1024),
}

# ── load summary ───────────────────────────────────────────────────────────────
df = pd.read_csv(os.path.join(RES_DIR, "summary.csv"))
df["group"] = df["model"].apply(
    lambda m: "TCGA" if m in TCGA_SET else ("Baseline" if m=="resnet50" else "Private"))
df["display"] = df["model"].map(DISP).fillna(df["model"])
df_sorted = df.sort_values("auc_macro_mean", ascending=False).reset_index(drop=True)

# ── load per-fold metrics ──────────────────────────────────────────────────────
fold_records = []
for _, row in df.iterrows():
    mname = row["model"]
    fp = os.path.join(RES_DIR, mname, "fold_metrics.csv")
    if os.path.exists(fp):
        fm = pd.read_csv(fp)
        fm["model"] = mname
        fm["group"] = row["group"]
        fm["display"] = DISP.get(mname, mname)
        fold_records.append(fm)
fold_df = pd.concat(fold_records, ignore_index=True)

# ── load clinical data ─────────────────────────────────────────────────────────
clin_raw = pd.read_csv(CLIN_CSV, encoding="utf-8-sig")
# rename columns by position (garbled encoding)
col_map = {
    clin_raw.columns[4]:  "institution",
    clin_raw.columns[5]:  "age",
    clin_raw.columns[6]:  "sex",        # 1=Male 2=Female
    clin_raw.columns[7]:  "race",
    clin_raw.columns[9]:  "tumor_loc",
    clin_raw.columns[14]: "extra_ext",
    clin_raw.columns[16]: "ajcc_stage",
    clin_raw.columns[17]: "treatment",
    clin_raw.columns[20]: "os_event",
    clin_raw.columns[21]: "os_time",
    clin_raw.columns[22]: "dss_event",
    clin_raw.columns[23]: "dss_time",
    clin_raw.columns[36]: "histology",
    clin_raw.columns[43]: "mitotic_rate",
    clin_raw.columns[44]: "levm",       # looping extravascular matrix
}
clin = clin_raw.rename(columns=col_map).copy()
clin["scna"] = pd.to_numeric(clin["SCNA Cluster No."], errors="coerce")
clin["d3m3"] = clin["scna"].apply(
    lambda x: 0 if x in (1,2) else (1 if x in (3,4) else np.nan))
clin = clin[clin["d3m3"].notna()].copy()
clin["d3m3"] = clin["d3m3"].astype(int)
clin["age"]  = pd.to_numeric(clin["age"], errors="coerce")
clin["sex_label"] = clin["sex"].map({1:"Male", 2:"Female"})
clin["group_label"] = clin["d3m3"].map({0:"D3 (Disomy 3)", 1:"M3 (Monosomy 3)"})


# ══════════════════════════════════════════════════════════════════════════════
# HELPER: three-line table
# ══════════════════════════════════════════════════════════════════════════════
def save_three_line_table(tdf, title, fname, col_widths=None, fontsize=8):
    n_rows, n_cols = tdf.shape
    row_h  = 0.30
    head_h = 0.42
    top_margin = 0.30
    fig_h = top_margin + head_h + n_rows * row_h + 0.20
    fig_w = sum(col_widths) if col_widths else n_cols * 1.5

    fig, ax = plt.subplots(figsize=(fig_w, fig_h))
    ax.set_xlim(0, fig_w); ax.set_ylim(0, fig_h); ax.axis("off")

    if col_widths is None:
        col_widths = [fig_w / n_cols] * n_cols
    x_pos = [sum(col_widths[:i]) for i in range(n_cols)]

    top_y    = fig_h - top_margin
    header_y = top_y - head_h
    data_top = header_y - 0.06
    bottom_y = 0.12
    lw = 1.2

    ax.axhline(top_y,    color="black", linewidth=lw*1.5)
    ax.axhline(header_y, color="black", linewidth=lw)
    ax.axhline(bottom_y, color="black", linewidth=lw*1.5)

    # title above top rule
    ax.text(fig_w/2, top_y + 0.05, title,
            ha="center", va="bottom", fontsize=fontsize+1,
            fontweight="bold", fontfamily="Arial")

    # header row
    for ci, (col, cw, xp) in enumerate(zip(tdf.columns, col_widths, x_pos)):
        align = "left" if ci == 0 else "center"
        ax.text(xp + (0.06 if ci==0 else cw/2),
                (top_y + header_y)/2, col,
                ha=align, va="center", fontsize=fontsize,
                fontweight="bold", fontfamily="Arial")

    # data rows
    for ri, row in enumerate(tdf.itertuples(index=False)):
        y = data_top - ri * row_h
        if ri % 2 == 1:
            ax.add_patch(FancyBboxPatch(
                (0, y - row_h*0.82), fig_w, row_h*0.88,
                boxstyle="square,pad=0", linewidth=0, facecolor="#f5f5f5"))
        for ci, (val, cw, xp) in enumerate(zip(row, col_widths, x_pos)):
            align = "left" if ci == 0 else "center"
            style = "italic" if str(val).startswith("ResNet") else "normal"
            color = "#555" if style == "italic" else "black"
            ax.text(xp + (0.06 if ci==0 else cw/2),
                    y - row_h*0.32, str(val),
                    ha=align, va="center", fontsize=fontsize,
                    fontstyle=style, color=color, fontfamily="Arial")

    ax.set_ylim(0, fig_h + 0.05)
    plt.tight_layout(pad=0.1)
    fig.savefig(os.path.join(OUT_DIR, fname),
                dpi=DPI, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    print(f"  Saved: {fname}")

    # also export Excel
    xlsx_name = fname.replace(".png", ".xlsx")
    _save_excel_table(tdf, title, os.path.join(OUT_DIR, xlsx_name))


def _save_excel_table(tdf, title, xlsx_path):
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Table"
    thick = Side(style="medium", color="000000")
    thin  = Side(style="thin",   color="000000")
    none  = Side(style=None)
    fill_alt = PatternFill("solid", fgColor="F5F5F5")

    # title row
    ws.merge_cells(start_row=1, start_column=1,
                   end_row=1, end_column=len(tdf.columns))
    tc = ws.cell(row=1, column=1, value=title)
    tc.font      = Font(name="Arial", bold=True, size=11)
    tc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    tc.border    = Border(top=thick, left=none, right=none, bottom=none)
    ws.row_dimensions[1].height = 28

    # header row
    for ci, col in enumerate(tdf.columns, start=1):
        cell = ws.cell(row=2, column=ci, value=col)
        cell.font      = Font(name="Arial", bold=True, size=10)
        cell.alignment = Alignment(horizontal="center" if ci > 1 else "left",
                                   vertical="center", wrap_text=True)
        cell.border    = Border(top=thick, bottom=thin, left=none, right=none)
    ws.row_dimensions[2].height = 20

    # data rows
    n_rows = len(tdf)
    for ri, row in enumerate(tdf.itertuples(index=False), start=3):
        is_last = (ri == n_rows + 2)
        for ci, val in enumerate(row, start=1):
            cell = ws.cell(row=ri, column=ci, value=str(val))
            cell.font      = Font(name="Arial", size=9.5,
                                  italic=str(val).startswith("ResNet"))
            cell.alignment = Alignment(horizontal="center" if ci > 1 else "left",
                                       vertical="center", wrap_text=True)
            cell.border    = Border(top=none, bottom=(thick if is_last else none),
                                    left=none, right=none)
        if ri % 2 == 0:
            for ci in range(1, len(tdf.columns)+1):
                ws.cell(row=ri, column=ci).fill = fill_alt
        ws.row_dimensions[ri].height = 16

    for ci, col in enumerate(tdf.columns, start=1):
        max_len = max([len(str(col))] + [len(str(v)) for v in tdf.iloc[:, ci-1]])
        ws.column_dimensions[openpyxl.utils.get_column_letter(ci)].width = \
            min(max_len * 1.35 + 2, 42)

    wb.save(xlsx_path)
    print(f"  Saved: {os.path.basename(xlsx_path)}")


def savefig(fig, stem):
    for ext in ("pdf", "png"):
        fig.savefig(os.path.join(OUT_DIR, f"{stem}.{ext}"),
                    dpi=DPI, bbox_inches="tight", facecolor="white")
    plt.close(fig)
    print(f"  Saved: {stem}.pdf/.png")


# ══════════════════════════════════════════════════════════════════════════════
# TABLE 1 – Clinical Characteristics (D3 vs M3)
# ══════════════════════════════════════════════════════════════════════════════
def make_clinical_table():
    d3 = clin[clin["d3m3"] == 0]
    m3 = clin[clin["d3m3"] == 1]
    total = clin

    def fmt_n(val, n):
        return f"{val} ({val/n*100:.1f}%)"

    def fmt_median(series):
        s = pd.to_numeric(series, errors="coerce").dropna()
        return f"{s.median():.1f} ({s.min():.0f}–{s.max():.0f})"

    def cat_pval(col, v1, v2):
        from scipy.stats import chi2_contingency
        cats = list(set(v1.dropna().tolist() + v2.dropna().tolist()))
        ct = [[sum(v1==c) for c in cats], [sum(v2==c) for c in cats]]
        try:
            _, p, _, _ = chi2_contingency(ct)
            return f"{p:.3f}"
        except:
            return "—"

    def cont_pval(s1, s2):
        from scipy.stats import mannwhitneyu
        a = pd.to_numeric(s1, errors="coerce").dropna()
        b = pd.to_numeric(s2, errors="coerce").dropna()
        try:
            _, p = mannwhitneyu(a, b, alternative="two-sided")
            return f"{p:.3f}"
        except:
            return "—"

    nd, nm, nt = len(d3), len(m3), len(total)

    rows = []
    rows.append({"Characteristic": "n", "Total (n=80)": str(nt),
                 "D3 (n=38)": str(nd), "M3 (n=42)": str(nm), "p-value": ""})
    rows.append({"Characteristic": "Age, median (range)",
                 "Total (n=80)": fmt_median(total["age"]),
                 "D3 (n=38)": fmt_median(d3["age"]),
                 "M3 (n=42)": fmt_median(m3["age"]),
                 "p-value": cont_pval(d3["age"], m3["age"])})
    rows.append({"Characteristic": "Sex",
                 "Total (n=80)":"", "D3 (n=38)":"", "M3 (n=42)":"", "p-value":
                 cat_pval("sex_label", d3["sex_label"], m3["sex_label"])})
    rows.append({"Characteristic": "  Male",
                 "Total (n=80)": fmt_n(sum(total["sex"]==1), nt),
                 "D3 (n=38)": fmt_n(sum(d3["sex"]==1), nd),
                 "M3 (n=42)": fmt_n(sum(m3["sex"]==1), nm), "p-value": ""})
    rows.append({"Characteristic": "  Female",
                 "Total (n=80)": fmt_n(sum(total["sex"]==2), nt),
                 "D3 (n=38)": fmt_n(sum(d3["sex"]==2), nd),
                 "M3 (n=42)": fmt_n(sum(m3["sex"]==2), nm), "p-value": ""})

    # AJCC Stage
    rows.append({"Characteristic": "AJCC Stage",
                 "Total (n=80)":"", "D3 (n=38)":"", "M3 (n=42)":"",
                 "p-value": cat_pval("ajcc_stage", d3["ajcc_stage"], m3["ajcc_stage"])})
    for stage in sorted(total["ajcc_stage"].dropna().unique()):
        rows.append({"Characteristic": f"  {stage}",
                     "Total (n=80)": fmt_n(sum(total["ajcc_stage"]==stage), nt),
                     "D3 (n=38)": fmt_n(sum(d3["ajcc_stage"]==stage), nd),
                     "M3 (n=42)": fmt_n(sum(m3["ajcc_stage"]==stage), nm),
                     "p-value": ""})

    # Histology
    rows.append({"Characteristic": "Histological Type",
                 "Total (n=80)":"", "D3 (n=38)":"", "M3 (n=42)":"",
                 "p-value": cat_pval("histology", d3["histology"], m3["histology"])})
    for h in ["Spindle Cell", "Epithelioid Cell", "Mixed Cell Type"]:
        n_tot = sum(total["histology"].fillna("").str.contains(h.split()[0], case=False))
        n_d   = sum(d3["histology"].fillna("").str.contains(h.split()[0], case=False))
        n_m   = sum(m3["histology"].fillna("").str.contains(h.split()[0], case=False))
        rows.append({"Characteristic": f"  {h}",
                     "Total (n=80)": fmt_n(n_tot, nt),
                     "D3 (n=38)": fmt_n(n_d, nd),
                     "M3 (n=42)": fmt_n(n_m, nm),
                     "p-value": ""})

    # Treatment
    rows.append({"Characteristic": "Treatment",
                 "Total (n=80)":"", "D3 (n=38)":"", "M3 (n=42)":"",
                 "p-value": cat_pval("treatment", d3["treatment"], m3["treatment"])})
    for tx in sorted(total["treatment"].dropna().unique()):
        rows.append({"Characteristic": f"  {tx}",
                     "Total (n=80)": fmt_n(sum(total["treatment"]==tx), nt),
                     "D3 (n=38)": fmt_n(sum(d3["treatment"]==tx), nd),
                     "M3 (n=42)": fmt_n(sum(m3["treatment"]==tx), nm),
                     "p-value": ""})

    # OS
    rows.append({"Characteristic": "Overall Survival (events)",
                 "Total (n=80)": fmt_n(int(pd.to_numeric(total["os_event"],errors="coerce").sum()), nt),
                 "D3 (n=38)":   fmt_n(int(pd.to_numeric(d3["os_event"],errors="coerce").sum()), nd),
                 "M3 (n=42)":   fmt_n(int(pd.to_numeric(m3["os_event"],errors="coerce").sum()), nm),
                 "p-value": cont_pval(d3["os_time"], m3["os_time"])})
    rows.append({"Characteristic": "OS Time, median days (range)",
                 "Total (n=80)": fmt_median(total["os_time"]),
                 "D3 (n=38)": fmt_median(d3["os_time"]),
                 "M3 (n=42)": fmt_median(m3["os_time"]),
                 "p-value": "—"})

    # SCNA Cluster
    rows.append({"Characteristic": "SCNA Cluster",
                 "Total (n=80)":"", "D3 (n=38)":"", "M3 (n=42)":"", "p-value":""})
    for c in [1,2,3,4]:
        label = f"  Cluster {c} ({'D3' if c<=2 else 'M3'})"
        rows.append({"Characteristic": label,
                     "Total (n=80)": fmt_n(sum(total["scna"]==c), nt),
                     "D3 (n=38)": fmt_n(sum(d3["scna"]==c), nd) if c<=2 else "—",
                     "M3 (n=42)": fmt_n(sum(m3["scna"]==c), nm) if c>=3 else "—",
                     "p-value": ""})

    tdf = pd.DataFrame(rows)
    save_three_line_table(
        tdf,
        title="Table 1. Baseline Clinical Characteristics of the TCGA-UVM Cohort (n=80)",
        fname="table1_clinical_characteristics.png",
        col_widths=[3.8, 2.1, 2.1, 2.1, 1.4],
        fontsize=8,
    )


# ══════════════════════════════════════════════════════════════════════════════
# TABLE 2 – Model Overview
# ══════════════════════════════════════════════════════════════════════════════
def make_model_overview_table():
    rows = []
    for mname in df_sorted["model"]:
        grp = df_sorted.loc[df_sorted["model"]==mname,"group"].values[0]
        auc = df_sorted.loc[df_sorted["model"]==mname,"auc_macro_mean"].values[0]
        meta = MODEL_META.get(mname, (DISP.get(mname,mname),"—","—","—","—","—"))
        disp, inst, para, arch, params, dfeat = meta
        rows.append({
            "Model": disp,
            "Institution": inst,
            "Paradigm": para,
            "Architecture": arch,
            "Params (M)": f"~{params}M" if isinstance(params,int) else str(params),
            "Feature Dim": str(dfeat),
            "Pretraining Data": "TCGA" if grp=="TCGA" else ("ImageNet" if mname=="resnet50" else "Private"),
            "AUC (mean)": f"{auc:.4f}",
        })
    tdf = pd.DataFrame(rows)
    save_three_line_table(
        tdf,
        title="Table 2. Overview of 21 Models Evaluated in the UVM D3/M3 Benchmark",
        fname="table2_model_overview.png",
        col_widths=[2.1,1.5,1.4,1.6,1.3,1.2,1.7,1.3],
        fontsize=7.5,
    )


# ══════════════════════════════════════════════════════════════════════════════
# FIG S1 – Study Design / Pipeline Schematic
# ══════════════════════════════════════════════════════════════════════════════
def make_pipeline_figure():
    fig, ax = plt.subplots(figsize=(13, 4.5))
    ax.set_xlim(0, 13); ax.set_ylim(0, 4.5); ax.axis("off")

    def box(x, y, w, h, color, text, fontsize=9, text_color="white", radius=0.18):
        ax.add_patch(FancyBboxPatch((x, y), w, h, boxstyle=f"round,pad={radius}",
                                    facecolor=color, edgecolor="white",
                                    linewidth=1.5, zorder=3))
        ax.text(x+w/2, y+h/2, text, ha="center", va="center",
                fontsize=fontsize, color=text_color, fontweight="bold",
                zorder=4, wrap=True)

    def arrow(x1, x2, y, color="#555", lw=2):
        ax.annotate("", xy=(x2, y), xytext=(x1, y),
                    arrowprops=dict(arrowstyle="-|>", color=color,
                                   lw=lw, mutation_scale=14), zorder=2)

    # Step 1: TCGA-UVM Cohort
    box(0.2, 2.2, 2.1, 1.5, "#3A6186",
        "TCGA-UVM Cohort\nn = 80 WSIs\nD3: 38 | M3: 42", fontsize=8.5)

    arrow(2.35, 2.85, 2.95)

    # Step 2: Patch Extraction
    box(2.85, 2.2, 2.1, 1.5, "#1A936F",
        "Trident Patch Extraction\n20× magnification\n256×256 px, 0 px overlap", fontsize=8)

    arrow(5.0, 5.5, 2.95)

    # Step 3: Feature Extraction (21 models)
    box(5.5, 2.0, 2.3, 1.9, "#9B2335",
        "Feature Extraction\n21 Pathology FMs\n+ ResNet-50 baseline\n(frozen encoder)", fontsize=8)

    arrow(7.85, 8.35, 2.95)

    # Step 4: ABMIL
    box(8.35, 2.2, 2.0, 1.5, "#5C4E8E",
        "ABMIL Aggregation\nn_token = 1\n(slide-level repr.)", fontsize=8)

    arrow(10.4, 10.9, 2.95)

    # Step 5: 5-fold CV + Evaluation
    box(10.9, 2.2, 1.85, 1.5, "#B06A00",
        "5-Fold Stratified CV\nAUC / Acc / F1", fontsize=8)

    # Label boxes below arrows
    for x, label in [
        (0.2+1.05, "80 cases"),
        (2.85+1.05, "~4,000–30,000\npatches/WSI"),
        (5.5+1.15, "D_feat: 384–2560"),
        (8.35+1.0, "Attention pooling"),
    ]:
        ax.text(x, 2.05, label, ha="center", va="top",
                fontsize=7.5, color="#444", style="italic")

    # Annotation: 3 scientific questions (bottom)
    for xi, (q_num, q_label, q_color) in enumerate([
        ("Q1", "Transfer Value\n(FM vs ResNet-50)", "#2E86AB"),
        ("Q2", "Pretraining Overlap\n(TCGA vs Private)", "#E07B39"),
        ("Q3", "Kaiko 2×2 Factorial\n(Capacity × Resolution)", "#9B59B6"),
    ]):
        bx = 1.5 + xi * 3.8
        ax.add_patch(FancyBboxPatch((bx, 0.1), 3.3, 1.0,
                                    boxstyle="round,pad=0.1",
                                    facecolor=q_color, alpha=0.15,
                                    edgecolor=q_color, linewidth=1.2, zorder=1))
        ax.text(bx+1.65, 0.62, f"Scientific Q{q_num[-1]}", ha="center",
                va="center", fontsize=8.5, fontweight="bold", color=q_color)
        ax.text(bx+1.65, 0.28, q_label, ha="center", va="center",
                fontsize=7.5, color=q_color)

    ax.set_title("Study Design and Computational Pipeline",
                 fontsize=12, fontweight="bold", pad=8)
    savefig(fig, "figS1_pipeline")


# ══════════════════════════════════════════════════════════════════════════════
# FIG 2 – Per-fold AUC Box Plot (all 21 models)
# ══════════════════════════════════════════════════════════════════════════════
def make_boxplot_pf():
    order = df_sorted["model"].tolist()
    order_disp = [DISP.get(m, m) for m in order]
    group_colors = {m: (C_TCGA if g=="TCGA" else (C_BASELINE if g=="Baseline" else C_PRIVATE))
                    for m, g in zip(df_sorted["model"], df_sorted["group"])}

    fig, ax = plt.subplots(figsize=(5.5, 10))
    positions = list(range(len(order)))

    for i, mname in enumerate(order):
        vals = fold_df[fold_df["model"] == mname]["auc_macro"].values
        color = group_colors[mname]
        bp = ax.boxplot(vals, positions=[i], vert=False, widths=0.5,
                        patch_artist=True, notch=False,
                        medianprops=dict(color="white", linewidth=2),
                        whiskerprops=dict(linewidth=1.0, color=color),
                        capprops=dict(linewidth=1.0, color=color),
                        flierprops=dict(marker="", linewidth=0))
        bp["boxes"][0].set_facecolor(color)
        bp["boxes"][0].set_alpha(0.65)
        bp["boxes"][0].set_linewidth(0.8)
        bp["boxes"][0].set_edgecolor(color)

        # jitter strip
        rng = np.random.default_rng(i)
        jitter = rng.uniform(-0.18, 0.18, len(vals))
        ax.scatter(vals, i + jitter, color=color, s=22, zorder=5,
                   edgecolors="white", linewidths=0.4, alpha=0.9)

    # ResNet-50 dashed line
    resnet_auc = df_sorted.loc[df_sorted["model"]=="resnet50","auc_macro_mean"].values[0]
    ax.axvline(resnet_auc, color=C_BASELINE, linewidth=1.2, linestyle="--", zorder=0,
               label=f"ResNet-50 mean ({resnet_auc:.4f})")

    ax.set_yticks(positions)
    ax.set_yticklabels(order_disp, fontsize=8)
    ax.set_xlabel("AUC (5-fold cross-validation)", fontsize=10)
    ax.set_title("Per-Fold AUC Distribution — All 21 Models\n"
                 "UVM D3/M3 Classification  |  n = 80  |  ABMIL",
                 fontsize=10, fontweight="bold")
    ax.set_xlim(0.60, 1.08)
    ax.xaxis.grid(True, color="#e0e0e0", linewidth=0.6)
    ax.set_axisbelow(True)

    legend_handles = [
        mpatches.Patch(facecolor=C_PRIVATE, alpha=0.7, label="Private-Pretrained"),
        mpatches.Patch(facecolor=C_TCGA,    alpha=0.7, label="TCGA-Pretrained"),
        mpatches.Patch(facecolor=C_BASELINE,alpha=0.7, label="Baseline (ResNet-50)"),
        Line2D([0],[0], color=C_BASELINE, lw=1.2, linestyle="--",
               label=f"ResNet-50 mean"),
    ]
    ax.legend(handles=legend_handles, loc="lower right", fontsize=8,
              frameon=True, framealpha=0.9, edgecolor="#ccc")

    plt.tight_layout()
    savefig(fig, "figS2_boxplot_per_fold_auc")


# ══════════════════════════════════════════════════════════════════════════════
# FIG 3 – Averaged ROC Curves (top 5 + ResNet50)
# ══════════════════════════════════════════════════════════════════════════════
def make_roc_curves():
    top5 = df_sorted[df_sorted["model"] != "resnet50"]["model"].head(5).tolist()
    show_models = top5 + ["resnet50"]
    colors = ["#E63946","#2A9D8F","#9B59B6","#E07B39","#1A936F","#888888"]

    fig, ax = plt.subplots(figsize=(5.5, 5.2))
    ax.plot([0,1],[0,1], color="#cccccc", linewidth=1.0, linestyle="--",
            label="Random classifier (AUC=0.50)")

    for mname, color in zip(show_models, colors):
        all_fpr, all_tpr = [], []
        base_fpr = np.linspace(0, 1, 101)
        fold_aucs = []
        for fold in range(1, 6):
            pred_path = os.path.join(RES_DIR, mname, f"fold_{fold}", "predictions.csv")
            if not os.path.exists(pred_path):
                continue
            preds = pd.read_csv(pred_path)
            if "prob_class1" not in preds.columns:
                continue
            y_true = preds["label"].values
            y_score = preds["prob_class1"].values
            if len(np.unique(y_true)) < 2:
                continue
            fpr, tpr, _ = roc_curve(y_true, y_score)
            roc_auc = sk_auc(fpr, tpr)
            fold_aucs.append(roc_auc)
            tpr_interp = np.interp(base_fpr, fpr, tpr)
            tpr_interp[0] = 0.0
            all_tpr.append(tpr_interp)

        if not all_tpr:
            continue
        mean_tpr = np.mean(all_tpr, axis=0)
        std_tpr  = np.std(all_tpr, axis=0)
        mean_auc = np.mean(fold_aucs)
        std_auc  = np.std(fold_aucs)
        disp_name = DISP.get(mname, mname)

        ax.plot(base_fpr, mean_tpr, color=color, linewidth=2.0,
                label=f"{disp_name}  AUC={mean_auc:.4f}±{std_auc:.4f}",
                alpha=0.9)
        ax.fill_between(base_fpr, mean_tpr-std_tpr, mean_tpr+std_tpr,
                        color=color, alpha=0.10)

    ax.set_xlabel("1 − Specificity (False Positive Rate)", fontsize=10)
    ax.set_ylabel("Sensitivity (True Positive Rate)", fontsize=10)
    ax.set_title("Mean ROC Curves (5-fold CV) — Top 5 Models + Baseline\n"
                 "UVM D3/M3 Binary Classification  |  n = 80",
                 fontsize=10, fontweight="bold")
    ax.legend(fontsize=7.8, loc="lower right", frameon=True,
              framealpha=0.92, edgecolor="#ccc")
    ax.set_xlim(-0.01, 1.01); ax.set_ylim(-0.01, 1.05)
    ax.xaxis.grid(True, color="#e0e0e0", linewidth=0.6)
    ax.yaxis.grid(True, color="#e0e0e0", linewidth=0.6)
    ax.set_axisbelow(True)
    plt.tight_layout()
    savefig(fig, "fig5_roc_curves")


# ══════════════════════════════════════════════════════════════════════════════
# FIG 4 – AUC Stability: Mean vs Std (bubble chart)
# ══════════════════════════════════════════════════════════════════════════════
def make_stability_scatter():
    fig, ax = plt.subplots(figsize=(7.5, 5.0))

    for _, row in df_sorted.iterrows():
        color = C_TCGA if row["group"]=="TCGA" else (C_BASELINE if row["group"]=="Baseline" else C_PRIVATE)
        ax.scatter(row["auc_macro_mean"], row["auc_macro_std"],
                   color=color, s=90, alpha=0.85, zorder=4,
                   edgecolors="white", linewidths=0.6)
        offset = (0.002, 0.001)
        if row["model"] in ["hoptimus1", "gigapath", "phikon_v2", "midnight12k",
                             "kaiko-vitb8", "lunit-vits8"]:
            ax.annotate(DISP.get(row["model"], row["model"]),
                        (row["auc_macro_mean"], row["auc_macro_std"]),
                        xytext=(row["auc_macro_mean"]+offset[0],
                                row["auc_macro_std"]+offset[1]),
                        fontsize=7.5, color=color, fontstyle="italic")

    # quadrant lines
    mean_auc_all = df_sorted["auc_macro_mean"].mean()
    std_auc_all  = df_sorted["auc_macro_std"].mean()
    ax.axvline(mean_auc_all, color="#aaa", lw=0.8, linestyle=":")
    ax.axhline(std_auc_all,  color="#aaa", lw=0.8, linestyle=":")

    ax.text(mean_auc_all + 0.001, df_sorted["auc_macro_std"].max() - 0.003,
            "High AUC\nHigh Variance", fontsize=7.5, color="#888", ha="left")
    ax.text(mean_auc_all + 0.001, df_sorted["auc_macro_std"].min() + 0.001,
            "High AUC\nLow Variance ✓", fontsize=7.5, color="#2A9D8F", ha="left",
            fontweight="bold")
    ax.text(df_sorted["auc_macro_mean"].min(), std_auc_all + 0.001,
            "Low AUC\nHigh Variance", fontsize=7.5, color="#888")

    ax.set_xlabel("Mean AUC (5-fold CV)", fontsize=10)
    ax.set_ylabel("Std of AUC (5-fold CV) — Instability", fontsize=10)
    ax.set_title("Model Performance vs. Stability\nMean AUC vs. Cross-Fold Variance",
                 fontsize=10, fontweight="bold")
    ax.xaxis.grid(True, color="#e0e0e0", linewidth=0.6)
    ax.yaxis.grid(True, color="#e0e0e0", linewidth=0.6)
    ax.set_axisbelow(True)

    legend_handles = [
        mpatches.Patch(facecolor=C_PRIVATE, alpha=0.8, label="Private-Pretrained"),
        mpatches.Patch(facecolor=C_TCGA,    alpha=0.8, label="TCGA-Pretrained"),
        mpatches.Patch(facecolor=C_BASELINE,alpha=0.8, label="Baseline"),
    ]
    ax.legend(handles=legend_handles, fontsize=8, frameon=True,
              framealpha=0.92, edgecolor="#ccc")
    plt.tight_layout()
    savefig(fig, "figS3_stability_scatter")


# ══════════════════════════════════════════════════════════════════════════════
# FIG 5 – Feature Dimension vs AUC (scatter)
# ══════════════════════════════════════════════════════════════════════════════
def make_feat_dim_scatter():
    fig, axes = plt.subplots(1, 2, figsize=(11, 4.5))

    for ax_idx, (x_col, x_label, log_x) in enumerate([
        ("D_feat", "Feature Dimension", True),
    ]):
        ax = axes[ax_idx] if ax_idx < 2 else axes[0]
        for _, row in df_sorted.iterrows():
            color = C_TCGA if row["group"]=="TCGA" else (C_BASELINE if row["group"]=="Baseline" else C_PRIVATE)
            ax.errorbar(row[x_col], row["auc_macro_mean"],
                        yerr=row["auc_macro_std"],
                        fmt="o", color=color, markersize=7, alpha=0.85,
                        capsize=2.5, capthick=0.8, linewidth=0.8,
                        ecolor=color, zorder=4, elinewidth=0.8)
        if log_x:
            ax.set_xscale("log")
        ax.set_xlabel(x_label, fontsize=10)
        ax.set_ylabel("AUC (mean ± std)", fontsize=10)
        ax.set_title(f"Feature Dimension vs. AUC", fontsize=10, fontweight="bold")
        ax.yaxis.grid(True, color="#e0e0e0", linewidth=0.6)
        ax.set_axisbelow(True)

        # correlation
        x_vals = df_sorted[x_col].values
        y_vals = df_sorted["auc_macro_mean"].values
        r, p = stats.spearmanr(x_vals, y_vals)
        ax.text(0.97, 0.04, f"Spearman r={r:.3f}, p={p:.3f}",
                transform=ax.transAxes, ha="right", va="bottom", fontsize=8,
                bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="#ccc", lw=0.8))

    # Second panel: params vs AUC
    ax2 = axes[1]
    params_list = []
    for _, row in df_sorted.iterrows():
        meta = MODEL_META.get(row["model"])
        params_list.append(meta[4] if meta else 25)

    for i, (_, row) in enumerate(df_sorted.iterrows()):
        color = C_TCGA if row["group"]=="TCGA" else (C_BASELINE if row["group"]=="Baseline" else C_PRIVATE)
        p = params_list[i]
        ax2.errorbar(p, row["auc_macro_mean"],
                     yerr=row["auc_macro_std"],
                     fmt="o", color=color, markersize=7, alpha=0.85,
                     capsize=2.5, capthick=0.8, linewidth=0.8, elinewidth=0.8,
                     ecolor=color, zorder=4)

    ax2.set_xscale("log")
    ax2.set_xlabel("Model Parameters (M, log scale)", fontsize=10)
    ax2.set_ylabel("AUC (mean ± std)", fontsize=10)
    ax2.set_title("Model Size vs. AUC", fontsize=10, fontweight="bold")
    ax2.yaxis.grid(True, color="#e0e0e0", linewidth=0.6)
    ax2.set_axisbelow(True)

    r2, p2 = stats.spearmanr(params_list, df_sorted["auc_macro_mean"].values)
    ax2.text(0.97, 0.04, f"Spearman r={r2:.3f}, p={p2:.3f}",
             transform=ax2.transAxes, ha="right", va="bottom", fontsize=8,
             bbox=dict(boxstyle="round,pad=0.3", fc="white", ec="#ccc", lw=0.8))

    legend_handles = [
        mpatches.Patch(facecolor=C_PRIVATE, alpha=0.8, label="Private-Pretrained"),
        mpatches.Patch(facecolor=C_TCGA,    alpha=0.8, label="TCGA-Pretrained"),
        mpatches.Patch(facecolor=C_BASELINE,alpha=0.8, label="Baseline"),
    ]
    for ax_ in axes:
        ax_.legend(handles=legend_handles, fontsize=7.5, frameon=True,
                   framealpha=0.9, edgecolor="#ccc", loc="upper left")

    fig.suptitle("Architectural Properties vs. Downstream AUC\n"
                 "UVM D3/M3 Classification  |  n = 80",
                 fontsize=11, fontweight="bold", y=1.01)
    plt.tight_layout()
    savefig(fig, "figS4_architecture_vs_auc")


# ══════════════════════════════════════════════════════════════════════════════
# FIG 6 – Metric Correlation Heatmap
# ══════════════════════════════════════════════════════════════════════════════
def make_metric_correlation():
    metrics = ["auc_macro_mean", "acc_mean", "f1_weighted_mean"]
    metric_labels = ["AUC", "Accuracy", "F1\n(weighted)"]
    M = df[metrics].values
    n = len(metrics)
    corr = np.corrcoef(M.T)

    fig, ax = plt.subplots(figsize=(5.0, 4.5))
    im = ax.imshow(corr, cmap="RdBu_r", vmin=-1, vmax=1, aspect="auto")

    for i in range(n):
        for j in range(n):
            ax.text(j, i, f"{corr[i,j]:.3f}", ha="center", va="center",
                    fontsize=11, fontweight="bold",
                    color="white" if abs(corr[i,j]) > 0.65 else "black")

    ax.set_xticks(range(n)); ax.set_yticks(range(n))
    ax.set_xticklabels(metric_labels, fontsize=9)
    ax.set_yticklabels(metric_labels, fontsize=9)
    ax.set_title("Inter-Metric Correlation\n(Pearson r, across 21 Models)",
                 fontsize=10, fontweight="bold")
    cbar = fig.colorbar(im, ax=ax, shrink=0.8, pad=0.03)
    cbar.set_label("Pearson r", fontsize=8.5)
    cbar.ax.tick_params(labelsize=8)
    plt.tight_layout()
    savefig(fig, "figS5_metric_correlation")


# ══════════════════════════════════════════════════════════════════════════════
# FIG 7 – Pretraining Paradigm Grouped Analysis (SSL vs VLP vs Supervised)
# ══════════════════════════════════════════════════════════════════════════════
def make_paradigm_analysis():
    paradigm_map = {m: MODEL_META[m][2] if m in MODEL_META else "SSL"
                    for m in df["model"]}
    df["paradigm"] = df["model"].map(paradigm_map)
    paradigms = ["SSL", "VLP", "Supervised"]
    para_colors = {"SSL": "#2E86AB", "VLP": "#9B59B6", "Supervised": "#888888"}

    fig, axes = plt.subplots(1, 2, figsize=(11, 5.0),
                             gridspec_kw={"width_ratios": [1.1, 1]})

    # Left: strip + box by paradigm
    ax = axes[0]
    positions = [1, 2, 3]
    rng = np.random.default_rng(99)
    for pos, para in zip(positions, paradigms):
        vals = df[df["paradigm"]==para]["auc_macro_mean"].values
        color = para_colors[para]
        bp = ax.boxplot(vals, positions=[pos], widths=0.35,
                        patch_artist=True, notch=False,
                        medianprops=dict(color="white", linewidth=2),
                        whiskerprops=dict(linewidth=1.1),
                        capprops=dict(linewidth=1.1),
                        flierprops=dict(marker="", linewidth=0))
        bp["boxes"][0].set_facecolor(color); bp["boxes"][0].set_alpha(0.5)
        bp["boxes"][0].set_linewidth(1.1); bp["boxes"][0].set_edgecolor(color)
        jitter = rng.uniform(-0.1, 0.1, len(vals))
        ax.scatter(pos+jitter, vals, color=color, s=50, zorder=5,
                   edgecolors="white", linewidths=0.5, alpha=0.9)
        ax.text(pos, np.mean(vals)-0.015, f"n={len(vals)}\nmean={np.mean(vals):.4f}",
                ha="center", va="top", fontsize=7.5, color=color, fontweight="bold")

    ax.set_xticks([1,2,3])
    ax.set_xticklabels(["Self-Supervised\n(SSL)", "Vision-Language\n(VLP)",
                         "Supervised\n(ImageNet)"], fontsize=9)
    ax.set_ylabel("AUC (5-fold CV, mean)", fontsize=10)
    ax.set_title("Pretraining Paradigm Comparison\n(descriptive; models are not i.i.d. samples)",
                 fontsize=10, fontweight="bold")
    ax.set_ylim(0.78, 0.98)
    ax.yaxis.grid(True, color="#e0e0e0", linewidth=0.6); ax.set_axisbelow(True)
    for pos, para in zip(positions, paradigms):
        ax.text(pos, 0.785, para_colors[para], color=para_colors[para], ha="center",
                fontsize=1)

    # Right: ranked dot plot colored by paradigm
    ax2 = axes[1]
    for i, row in df_sorted.iterrows():
        color = para_colors.get(df[df["model"]==row["model"]]["paradigm"].values[0], "#888")
        ax2.errorbar(row["auc_macro_mean"], i,
                     xerr=row["auc_macro_std"],
                     fmt="o", color=color, markersize=5,
                     capsize=2.5, capthick=0.8, linewidth=0.8,
                     ecolor=color, alpha=0.85)
    ax2.set_yticks(range(len(df_sorted)))
    ax2.set_yticklabels(df_sorted["display"].tolist(), fontsize=7.5)
    ax2.set_xlabel("AUC (mean ± std)", fontsize=9)
    ax2.set_title("All Models by Paradigm", fontsize=9, fontweight="bold")
    ax2.xaxis.grid(True, color="#e0e0e0", linewidth=0.6); ax2.set_axisbelow(True)

    legend_handles = [
        mpatches.Patch(facecolor=para_colors[p], alpha=0.8, label=p)
        for p in paradigms
    ]
    ax2.legend(handles=legend_handles, fontsize=8, frameon=True,
               framealpha=0.9, edgecolor="#ccc", loc="lower right")

    fig.suptitle("Pretraining Paradigm Analysis — SSL vs. VLP vs. Supervised",
                 fontsize=11, fontweight="bold", y=1.01)
    plt.tight_layout()
    savefig(fig, "figS6_paradigm_analysis")


# ══════════════════════════════════════════════════════════════════════════════
# FIG 8 – Clinical Distribution: D3 vs M3
# ══════════════════════════════════════════════════════════════════════════════
def make_clinical_figure():
    fig, axes = plt.subplots(2, 3, figsize=(13, 8))
    fig.suptitle("Clinical Characteristics of TCGA-UVM Cohort — D3 vs M3 Comparison\n"
                 "n = 80 (D3: 38, M3: 42)",
                 fontsize=11.5, fontweight="bold", y=1.01)

    d3 = clin[clin["d3m3"]==0]
    m3 = clin[clin["d3m3"]==1]

    # (1) Age distribution
    ax = axes[0, 0]
    bins = np.linspace(20, 85, 14)
    ax.hist(d3["age"].dropna(), bins=bins, color=C_D3, alpha=0.7,
            label=f"D3 (n={len(d3)})", edgecolor="white", linewidth=0.5)
    ax.hist(m3["age"].dropna(), bins=bins, color=C_M3, alpha=0.7,
            label=f"M3 (n={len(m3)})", edgecolor="white", linewidth=0.5)
    _, p_age = stats.mannwhitneyu(d3["age"].dropna(), m3["age"].dropna(), alternative="two-sided")
    ax.set_xlabel("Age (years)"); ax.set_ylabel("Count")
    ax.set_title(f"Age Distribution  (p={p_age:.3f})", fontweight="bold")
    ax.legend(fontsize=8.5); ax.yaxis.grid(True, color="#e0e0e0", lw=0.6)
    ax.set_axisbelow(True)

    # (2) Sex
    ax = axes[0, 1]
    sex_d3 = [sum(d3["sex"]==1), sum(d3["sex"]==2)]
    sex_m3 = [sum(m3["sex"]==1), sum(m3["sex"]==2)]
    x = np.array([0, 1])
    w = 0.32
    ax.bar(x-w/2, sex_d3, width=w, color=C_D3, alpha=0.8, label="D3", edgecolor="white")
    ax.bar(x+w/2, sex_m3, width=w, color=C_M3, alpha=0.8, label="M3", edgecolor="white")
    ax.set_xticks([0,1]); ax.set_xticklabels(["Male","Female"])
    ax.set_ylabel("Count"); ax.set_title("Sex Distribution", fontweight="bold")
    ax.legend(fontsize=8.5); ax.yaxis.grid(True, color="#e0e0e0", lw=0.6)
    ax.set_axisbelow(True)

    # (3) SCNA Cluster
    ax = axes[0, 2]
    scna_vals = [1, 2, 3, 4]
    scna_n = [sum(clin["scna"]==c) for c in scna_vals]
    scna_colors = [C_D3, C_D3, C_M3, C_M3]
    bars = ax.bar(scna_vals, scna_n, color=scna_colors, alpha=0.8, edgecolor="white")
    for bar, n in zip(bars, scna_n):
        ax.text(bar.get_x()+bar.get_width()/2, bar.get_height()+0.3,
                str(n), ha="center", va="bottom", fontsize=9, fontweight="bold")
    ax.set_xticks(scna_vals)
    ax.set_xticklabels(["Cluster 1\n(D3)","Cluster 2\n(D3)",
                         "Cluster 3\n(M3)","Cluster 4\n(M3)"])
    ax.set_ylabel("Count"); ax.set_title("SCNA Cluster Distribution", fontweight="bold")
    ax.yaxis.grid(True, color="#e0e0e0", lw=0.6); ax.set_axisbelow(True)

    # (4) AJCC Stage
    ax = axes[1, 0]
    stages = sorted(clin["ajcc_stage"].dropna().unique())
    d3_s = [sum(d3["ajcc_stage"]==s) for s in stages]
    m3_s = [sum(m3["ajcc_stage"]==s) for s in stages]
    x = np.arange(len(stages)); w = 0.35
    ax.bar(x-w/2, d3_s, width=w, color=C_D3, alpha=0.8, label="D3", edgecolor="white")
    ax.bar(x+w/2, m3_s, width=w, color=C_M3, alpha=0.8, label="M3", edgecolor="white")
    ax.set_xticks(x); ax.set_xticklabels(stages, rotation=20, ha="right", fontsize=8)
    ax.set_ylabel("Count"); ax.set_title("AJCC Stage", fontweight="bold")
    ax.legend(fontsize=8.5); ax.yaxis.grid(True, color="#e0e0e0", lw=0.6)
    ax.set_axisbelow(True)

    # (5) Histological type
    ax = axes[1, 1]
    hist_cats = ["Spindle", "Epithelioid", "Mixed"]
    d3_h = [sum(d3["histology"].fillna("").str.contains(h, case=False)) for h in hist_cats]
    m3_h = [sum(m3["histology"].fillna("").str.contains(h, case=False)) for h in hist_cats]
    x = np.arange(len(hist_cats)); w = 0.35
    ax.bar(x-w/2, d3_h, width=w, color=C_D3, alpha=0.8, label="D3", edgecolor="white")
    ax.bar(x+w/2, m3_h, width=w, color=C_M3, alpha=0.8, label="M3", edgecolor="white")
    ax.set_xticks(x); ax.set_xticklabels(["Spindle Cell","Epithelioid Cell","Mixed"], fontsize=8.5)
    ax.set_ylabel("Count"); ax.set_title("Histological Type", fontweight="bold")
    ax.legend(fontsize=8.5); ax.yaxis.grid(True, color="#e0e0e0", lw=0.6)
    ax.set_axisbelow(True)

    # (6) OS Time distribution
    ax = axes[1, 2]
    d3_ost = pd.to_numeric(d3["os_time"], errors="coerce").dropna()
    m3_ost = pd.to_numeric(m3["os_time"], errors="coerce").dropna()
    bins2  = np.linspace(0, max(d3_ost.max(), m3_ost.max())+50, 12)
    ax.hist(d3_ost, bins=bins2, color=C_D3, alpha=0.7,
            label=f"D3 (n={len(d3_ost)})", edgecolor="white", linewidth=0.5)
    ax.hist(m3_ost, bins=bins2, color=C_M3, alpha=0.7,
            label=f"M3 (n={len(m3_ost)})", edgecolor="white", linewidth=0.5)
    _, p_os = stats.mannwhitneyu(d3_ost, m3_ost, alternative="two-sided")
    ax.set_xlabel("Overall Survival Time (days)"); ax.set_ylabel("Count")
    ax.set_title(f"OS Time Distribution  (p={p_os:.3f})", fontweight="bold")
    ax.legend(fontsize=8.5); ax.yaxis.grid(True, color="#e0e0e0", lw=0.6)
    ax.set_axisbelow(True)

    plt.tight_layout()
    savefig(fig, "fig1_clinical_characteristics")


# ══════════════════════════════════════════════════════════════════════════════
# FIG 9 – Multi-metric Radar / Bar (all metrics side by side for top 10)
# ══════════════════════════════════════════════════════════════════════════════
def make_multimetric_bar():
    top10 = df_sorted.head(10)
    metrics = ["auc_macro_mean","acc_mean","f1_weighted_mean"]
    metric_labels = ["AUC","Accuracy","F1 (weighted)"]
    n = len(top10)
    x = np.arange(n)
    w = 0.22

    fig, ax = plt.subplots(figsize=(12, 5.0))
    palette = ["#2E86AB","#1A936F","#E07B39"]
    for mi, (metric, label, color) in enumerate(zip(metrics, metric_labels, palette)):
        offset = (mi - 1) * w
        bars = ax.bar(x + offset, top10[metric], width=w,
                      color=color, alpha=0.82, label=label, edgecolor="white",
                      linewidth=0.5)

    ax.set_xticks(x)
    ax.set_xticklabels(top10["display"], rotation=28, ha="right", fontsize=8.5)
    ax.set_ylabel("Score", fontsize=10)
    ax.set_ylim(0, 1.05)
    ax.set_title("Multi-Metric Comparison — Top 10 Foundation Models\n"
                 "UVM D3/M3 Classification  |  n = 80",
                 fontsize=10.5, fontweight="bold")
    ax.yaxis.grid(True, color="#e0e0e0", linewidth=0.6); ax.set_axisbelow(True)
    ax.legend(fontsize=9, frameon=True, framealpha=0.9, edgecolor="#ccc",
              loc="upper right")

    # Add group labels (TCGA / Private) below model names
    for i, (_, row) in enumerate(top10.iterrows()):
        grp_lbl = "★TCGA" if row["group"]=="TCGA" else ""
        if grp_lbl:
            ax.text(i, -0.07, grp_lbl, ha="center", va="top",
                    fontsize=7, color=C_TCGA, transform=ax.get_xaxis_transform())

    plt.tight_layout()
    savefig(fig, "figS7_multimetric_top10")


# ══════════════════════════════════════════════════════════════════════════════
# MAIN
# ══════════════════════════════════════════════════════════════════════════════
if __name__ == "__main__":
    print(f"Output: {OUT_DIR}\n")

    print("[Table 1] Clinical characteristics ...")
    make_clinical_table()

    print("[Table 2] Model overview ...")
    make_model_overview_table()

    print("[Fig S1] Pipeline schematic ...")
    make_pipeline_figure()

    print("[Fig 2] Per-fold box plots ...")
    make_boxplot_pf()

    print("[Fig 3] ROC curves ...")
    make_roc_curves()

    print("[Fig 4] Stability scatter ...")
    make_stability_scatter()

    print("[Fig 5] Architecture vs AUC ...")
    make_feat_dim_scatter()

    print("[Fig 6] Metric correlation ...")
    make_metric_correlation()

    print("[Fig 7] Paradigm analysis ...")
    make_paradigm_analysis()

    print("[Fig 8] Clinical distribution ...")
    make_clinical_figure()

    print("[Fig 9] Multi-metric bar ...")
    make_multimetric_bar()

    print("\nAll done. Figures saved to:", OUT_DIR)
